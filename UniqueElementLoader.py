from openpyxl import load_workbook

#Load what colums you want to check
lines = [
    "Product category",
    "Colour",
    "Material",
    "Finish",
    "Edge material",
    "Edge colour",
    "Fabric - composition",
    "Fabric - colour",
    "Handle material",
    "Handle colour",
    "Type of legs",
    "Legs colour"
]
skipValues = [
    "None",
    "N/A",
    "Null",
    "",
    None
]

file_path = input("Drag your file here: ")
print("")
file = load_workbook(file_path).active
indexes = {}
uniqueParametres = {}

#Loads indexes
for cell in file[1]:
    indexes[cell.value] = cell.column -1

#Loops thru the values, storing them in a "HashSet"
for row in file.iter_rows(values_only=True):
    for parametr in lines:
        value = str(row[indexes.get(parametr)]).strip()
    
        #Skips useless values
        if value == parametr or value in skipValues:
            continue
        
        value = value.lower()

        if parametr not in uniqueParametres:
            uniqueParametres[parametr] = set()
        uniqueParametres[parametr].add(value)

#Prints
for param, values in uniqueParametres.items():
    value_list = ", ".join(sorted(values))
    print(f"{param}: {value_list}")
    print()

input("Press enter to exit")